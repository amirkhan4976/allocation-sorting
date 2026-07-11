# Import required classes and functions from openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import range_boundaries

# -----------------------------
# Load the source Excel workbook
# -----------------------------
excel_file = "Allocation Sorting.xlsx"
wb = load_workbook(filename=excel_file)
work_sheet = wb.active

# ----------------------------------------------------
# Create a new workbook to store Workable and Skip data
# ----------------------------------------------------
wn_allocation_workbook = Workbook()

# Rename the default worksheet to "Workable"
wn_allocation_workbook_workable_worksheet = wn_allocation_workbook.active
wn_allocation_workbook_workable_worksheet.title = "Workable"

# Create another worksheet named "Skips"
wn_allocation_workbook_skip_worksheet = wn_allocation_workbook.create_sheet(title="Skips")


# -------------------------
# Configuration / Constants
# -------------------------
name_table_inventory = "Inventory"      # Inventory table name
name_table_skips = "Skips"              # Skips table name
header_check = "Account Number"         # Column used for comparison
insert_skip_work = "python_input"       # New column to indicate Work/Skip


# ------------------------------
# Get the Inventory Excel table
# ------------------------------
table_inventory = work_sheet.tables[name_table_inventory]

# Count the total number of columns in the Inventory table
total_columns = len(table_inventory.tableColumns)

# Insert a new worksheet column immediately after the table
work_sheet.insert_cols(total_columns + 1)

# Add a header for the new column
work_sheet.cell(
    row=1,
    column=total_columns + 1
).value = insert_skip_work


# ----------------------------------------------------
# Expand the Inventory table so it includes new column
# ----------------------------------------------------
start, end = table_inventory.ref.split(":")

# New ending column becomes one column larger
end_column = len(table_inventory.tableColumns) + 1

# Get the last row occupied by the table
_, _, _, last_row_of_table = range_boundaries(table_inventory.ref)

# Build the new ending cell reference
new_end = f"{get_column_letter(end_column)}{last_row_of_table}"

# Update the Inventory table range
table_inventory.ref = f"{start}:{new_end}"


# --------------------------
# Shift the Skips table also
# --------------------------
table_skips = work_sheet.tables[name_table_skips]

# Get the current table boundaries
min_col, min_row, max_col, max_row = range_boundaries(table_skips.ref)

# Move the entire table one column to the right because a new column
# was inserted before it
table_skips.ref = (
    f"{get_column_letter(min_col + 1)}{min_row}:"
    f"{get_column_letter(max_col + 1)}{max_row}"
)


# -----------------------------------------
# Read all rows from both Inventory & Skips
# -----------------------------------------
rows_inventory = work_sheet[table_inventory.ref]
rows_skips = work_sheet[table_skips.ref]


# ------------------------------------------------
# Store all account numbers from the Skips table
# ------------------------------------------------
skips_list = []

# Ignore the header row
for row in rows_skips[1:]:
    for cell in row:
        skips_list.append(cell.value)


# --------------------------------
# Read Inventory table header row
# --------------------------------
headers_inventory = [cell.value for cell in rows_inventory[0]]

# Copy headers to both new worksheets
wn_allocation_workbook_workable_worksheet.append(headers_inventory)
wn_allocation_workbook_skip_worksheet.append(headers_inventory)


# ------------------------------------------------
# Find the required column indexes dynamically
# ------------------------------------------------
col_header_check_inventory = headers_inventory.index(header_check)
col_insert_skip_work = headers_inventory.index(insert_skip_work)


# ------------------------
# Counters for final total
# ------------------------
skips_count = 0
workable = 0


# ---------------------------------------------
# Process every Inventory record (except header)
# ---------------------------------------------
for row in rows_inventory[1:]:

    # Read Account Number
    check_account = row[col_header_check_inventory].value

    # If account exists in the Skips list
    if check_account in skips_list:

        # Mark it as Skip in the original workbook
        row[col_insert_skip_work].value = "Skip"

        # Copy row into Skips worksheet
        # Replace None values with "-"
        wn_allocation_workbook_skip_worksheet.append(
            ["-" if cell.value is None else cell.value for cell in row]
        )

        skips_count += 1

    # Otherwise mark it as Work
    else:

        row[col_insert_skip_work].value = "Work"

        wn_allocation_workbook_workable_worksheet.append(
            ["-" if cell.value is None else cell.value for cell in row]
        )

        workable += 1


# ----------------------
# Display final counts
# ----------------------
print(f"Total skips are: {skips_count} and workable are: {workable}")


# ----------------------------------------------------
# Save the modified original workbook with new column
# ----------------------------------------------------
try:
    wb.save(f"New_{excel_file}")

    print(f"File name: New_{excel_file}")
    print("[+] Changes saved successfully")

except Exception as e:
    print(e)
    print("[-] Failed to save changes.")
    print("[-] Please try again by closing the file if opened.")


# --------------------------------------------
# Save the newly created allocation workbook
# --------------------------------------------
try:
    import datetime

    # Generate filename using today's date
    todays_date = datetime.date.today()
    allocation_file_name = f"Allocation-{todays_date}.xlsx"

    # Save the workbook
    wn_allocation_workbook.save(allocation_file_name)

    print("[+] New allocation file created...")
    print(f"File name: {allocation_file_name}")

except Exception as e:
    print(e)
    print("[-] Failed to create new allocation file.")